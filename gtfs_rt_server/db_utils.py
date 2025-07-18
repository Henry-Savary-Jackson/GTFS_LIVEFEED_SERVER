from sqlalchemy.dialects.sqlite import insert
from sqlalchemy import select
import pandas as pd
from flask import current_app
from pathlib import Path
import datetime
import openpyxl
from gtfs_rt_server.schema import (
    TripUpdate,
    Alert,
    EntityTypes,
    Causes,
    Effects,
    InformedEntityToAlerts,
    TripUpdateToStop,
    db,
    User,
    Role,
    roles_users,
    get_user_by_username,
)
from sqlalchemy import text, func
from typing import Optional
from argon2 import PasswordHasher
from flask import current_app


def get_password_hasher():
    return PasswordHasher()


password_hasher = get_password_hasher()


def get_route_id_of_trip(trip_id):
    with db.session.begin():
        return db.session.execute(
            text("SELECT route_id FROM trips WHERE trip_id = :trip_id"),
            {"trip_id": trip_id},
        ).fetchall()[0][0]


def add_role(role_name):
    with db.session.begin():
        db.session.execute(insert(Role).values(name=role_name).on_conflict_do_nothing())


def delete_user_with_username(username):
    user = get_user_by_username(username)
    if user is None:
        raise Exception(f"User '{username}' not found.")

    try:
        with db.session.begin():
            db.session.delete(user)
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        raise e


def list_users():
    with db.session.begin():
        return db.session.query(User).all()


def modify_user(id, username=None, password=None, roles=None):
    with db.session.begin():
        try:
            user = db.session.query(User).where(User.user_id == id).first()
            if not user:
                raise Exception(f"User with id '{id}' not found.")
            if username:
                user.username = username
            if password:
                user.password = password_hasher.hash(password)
            if roles is not None:
                user.roles = [
                    db.session.query(Role).filter(Role.name == name).first()
                    or Role(name=name)
                    for name in roles
                ]
            db.session.merge(user)

        except Exception as e:
            db.session.rollback()
            raise e


def insert_user(username, rawPassword, roles=[]):
    with db.session.begin():
        try:
            user = db.session.query(User).where(User.username == username).first()
            hash_pass = password_hasher.hash(rawPassword)
            roles = [
                (
                    db.session.query(Role).filter(Role.name == name).first()
                    or Role(name=name)
                )
                for name in roles
            ]
            if user:
                user.hash_pass = hash_pass
                user.roles = roles
                db.session.merge(user)
            else:
                user = User(
                    username=username,
                    hash_pass=hash_pass,
                    roles=roles,
                )
                db.session.add(user)
        except Exception as e:
            db.session.rollback()
            raise e


# add a column for active trips and inactive trips
def get_trips(service=None, route=None, number=None, time_after=None):
    with db.session.begin():
        sql = "SELECT DISTINCT stop_times.trip_id , trip_headsign AS endTerminus "
        if time_after:
            sql += " , CASE WHEN ( MAX(arrival_time) >= :time_after ) THEN ( CASE WHEN ( MIN(arrival_time) <= :time_after ) THEN 0 ELSE 1 END ) ELSE 2 END  AS inprogress "
        sql += " FROM stop_times INNER JOIN trips ON stop_times.trip_id = trips.trip_id"
        if service or route or number:
            sql += " WHERE "
        if service:
            sql += " service_id = :service "
            if route or number:
                sql += " AND "
        if route:
            sql += " route_id = :route "
            if number:
                sql += "AND "
        if number:
            sql += " trips.trip_id LIKE :tripid "
        if time_after:
            sql += " GROUP BY trips.trip_id ORDER BY inprogress "
        params = dict()
        params["service"] = service
        params["route"] = route
        params["tripid"] = f"%-{number}%"
        params["time_after"] = time_after
        trips = db.session.execute(text(sql), params=params).fetchall()
        return [trip._asdict() for trip in trips]


def get_stops(stop_name=None):
    with db.session.begin():
        sql = "SELECT stop_id, stop_name FROM stops "
        params = dict()
        if stop_name:
            sql += " WHERE stop_name LIKE :stop_name"
            params["stop_name"] = f"%{stop_name}%"

        stops = db.session.execute(text(sql), params).fetchall()
        return [list(stop) for stop in stops]


def get_trip_ids_routes():
    # create dict of trips to their route
    with db.session.begin():
        sql = "SELECT trip_id , route_id FROM trips "
        rows = db.session.execute(text(sql)).fetchall()
        return {row[0]: row[1] for row in rows}


def get_routes():
    with db.session.begin():
        sql = "SELECT route_id, route_long_name, route_color FROM routes "
        routes = db.session.execute(text(sql)).fetchall()
        return [route._asdict() for route in routes]


def get_services():
    with db.session.begin():
        services = db.session.execute(
            text("SELECT service_id FROM calendar ")
        ).fetchall()
        return [service[0] for service in services]


def get_number_of_stoptimes(trip_id):
    with db.session.begin():
        count = db.session.execute(
            text("SELECT COUNT(*) FROM stop_times WHERE trip_id = :trip_id"),
            {"trip_id": trip_id},
        ).fetchall()
        return count[0][0]


def get_stoptimes_of_trip(trip_id, include_time=True):
    # print(trip_id, "checking")
    with db.session.begin():
        stoptimes = db.session.execute(
            text(
                f"SELECT stop_sequence AS stopSequence, stop_id AS stopId {', strftime("%H:%M:%S", arrival_time) as arrival' if include_time else ''} FROM stop_times WHERE trip_id = :trip_id"
            ),
            {"trip_id": trip_id},
        ).fetchall()
        return [stoptime._asdict() for stoptime in stoptimes]


def route_exists(route_id):
    # print(route_id, "checking")
    with db.session.begin():
        result = db.session.execute(
            text("SELECT 1 FROM routes WHERE route_id = :route_id"),
            {"route_id": route_id},
        ).fetchall()
        return bool(result)


def stop_exists(stop_id):
    # print(stop_id, "checking")
    current_app.logger.debug(f"checking {stop_id}")
    with db.session.begin():
        result = db.session.execute(
            text("SELECT 1 FROM stops WHERE stop_id = :stop_id"), {"stop_id": stop_id}
        ).fetchall()
        return bool(result)


def stop_on_route(stop_id, route_id):
    current_app.logger.debug(f"checking {stop_id} {route_id}")
    # print(stop_id, route_id, "checking")
    with db.session.begin():
        result = db.session.execute(
            text(
                "SELECT 1 FROM stop_times INNER JOIN trips ON stop_times.trip_id = trips.trip_id WHERE stop_id = :stop_id AND route_id = :route_id"
            ),
            {"stop_id": stop_id, "route_id": route_id},
        ).fetchall()
        return bool(result)


def trip_exists(trip_id):
    with db.session.begin():
        result = db.session.execute(
            text("SELECT 1 FROM trips WHERE trip_id = :trip_id "), {"trip_id": trip_id}
        ).fetchall()
        return bool(result)


def agency_exists(agency_id):
    with db.session.begin():
        result = db.session.execute(
            text("SELECT 1 FROM agency WHERE agency_id = :agency_id "),
            {"agency_id": agency_id},
        ).fetchall()
        return bool(result)


# given an alert's id , delete it from the database
def delete_alert_from_log(alert_id):
    with db.session.begin():
        db.session.execute(db.delete(Alert).where(Alert.alert_id == alert_id))
        db.session.commit()


# given a tripUpdates's id , delete it from the database
def delete_trip_update_from_log(trip_update_id):
    with db.session.begin():
        db.session.execute(
            db.delete(TripUpdate).where(TripUpdate.trip_update_id == trip_update_id)
        )
        db.session.commit()


# given a json object of an alert, add or update into database
def add_alert_to_db(id, alert):

    alert_data = {
        "alert_id": id,
        "cause": getattr(Causes, alert["cause"]),
        "effect": getattr(Effects, alert["effect"]),
        "date": datetime.date.today().isoformat()
    }
    if "activePeriod" in alert and len(alert["activePeriod"]) > 0:
        if "start" in alert["activePeriod"][0]:
            alert_data.update({"start_time": alert["activePeriod"][0]["start"]})
        if "end" in alert["activePeriod"][0]:
            alert_data.update({"end_time": alert["activePeriod"][0]["end"]})

    stmt = (
        insert(Alert)
        .values(**alert_data)
        .on_conflict_do_update(index_elements=["alert_id"], set_=alert_data)
    )

    with db.session.begin():
        db.session.execute(stmt)
        db.session.query(InformedEntityToAlerts).filter_by(
            alert_id=alert_data["alert_id"]
        ).delete()

        for entity in alert.get("informedEntity", []):

            entity_type = EntityTypes.trips
            entity_id = ""
            if "trip" in entity:
                entity_id = entity["trip"]["tripId"]
            if "stopId" in entity:
                entity_type = EntityTypes.stops
                entity_id = entity["stopId"]
            if "routeId" in entity:
                entity_type = EntityTypes.routes
                entity_id = entity["routeId"]

            db.session.execute(
                insert(InformedEntityToAlerts)
                .values(
                    alert_id=alert_data["alert_id"],
                    entity_id=entity_id,
                    entity_type=entity_type,
                )
                .on_conflict_do_nothing()
            )


# given a json object of a trip update, add or update into database
def add_trip_update_to_db(id, trip_update):

    update_data = {
        "trip_update_id": id,
        "trip_id": trip_update["trip"]["tripId"],
        "date": datetime.date.today().isoformat(),
        "cancelled": trip_update.get("cancelled", False),
    }
    update_data["route_id"] = get_route_id_of_trip(update_data["trip_id"])
    stoptimes = get_stoptimes_of_trip(update_data["trip_id"], include_time=False)

    stmt = (
        insert(TripUpdate)
        .values(**update_data)
        .on_conflict_do_update(index_elements=["trip_update_id"], set_=update_data)
    )

    with db.session.begin():
        db.session.execute(stmt)
        db.session.query(TripUpdateToStop).filter_by(trip_update_id=id).delete()

        for stop_update in trip_update.get("stopTimeUpdate", []):
            db.session.add(
                TripUpdateToStop(
                    trip_update_id=id,
                    stop_id=stoptimes[stop_update["stopSequence"]]["stopId"],
                    delay=(
                        None
                        if "arrival" not in stop_update
                        else stop_update["arrival"].get("delay", None)
                    ),
                    skip=stop_update.get("skip", False),
                )
            )


## TODO change capitals


#  create a result as pandas dataframe, where you group the count of alerts by trips
def get_alerts_by_trips():
    result = (
        db.session.query(
            InformedEntityToAlerts.entity_id,
            func.count(InformedEntityToAlerts.alert_id).label("alert_count"),
        )
        .filter(InformedEntityToAlerts.entity_type == EntityTypes.trips)
        .group_by(InformedEntityToAlerts.entity_id)
        .all()
    )

    # return list(result)

    return pd.DataFrame(result, columns=["trip_id", "alert_count"])


#  create a result as pandas dataframe, where you group the count of alerts by route
def get_alerts_by_route():
    result = (
        db.session.query(
            InformedEntityToAlerts.entity_id,
            func.count(InformedEntityToAlerts.alert_id).label("alert_count"),
        )
        .filter(InformedEntityToAlerts.entity_type == EntityTypes.routes)
        .group_by(InformedEntityToAlerts.entity_id)
        .all()
    )

    # return list(result)
    return pd.DataFrame(result, columns=["route_id", "alert_count"])


#  create a result as pandas dataframe, where you group the count of alerts by stop
def get_alerts_by_stop():
    result = (
        db.session.query(
            InformedEntityToAlerts.entity_id,
            func.count(InformedEntityToAlerts.alert_id).label("alert_count"),
        )
        .filter(InformedEntityToAlerts.entity_type == EntityTypes.stops)
        .group_by(InformedEntityToAlerts.entity_id)
        .all()
    )
    # return list(result)
    return pd.DataFrame(result, columns=["stop_id", "alert_count"])


#


#  create a result as pandas dataframe, where you group the count of trip updates by trips
def get_trip_updates_by_trips():

    result = (
        db.session.query(
            TripUpdate.trip_id,
            func.count(TripUpdate.trip_update_id).label("update_count"),
        )
        .group_by(TripUpdate.trip_id)
        .all()
    )

    # return list(result)
    return pd.DataFrame(result, columns=["trip_id", "update_count"])


#  create a result as pandas dataframe, where you group the count of trip updates by routes
def get_trip_updates_by_stops():
    result = (
        db.session.query(
            TripUpdateToStop.stop_id,
            func.count(TripUpdateToStop.trip_update_id).label("update_count"),
        )
        .group_by(TripUpdateToStop.stop_id)
        .all()
    )

    # return list(result)
    return pd.DataFrame(result, columns=["stop_id", "update_count"])


#  create a result as pandas dataframe, where you group the count of trip updates by stops
def get_trip_updates_by_routes():
    result = (
        db.session.query(
            TripUpdate.route_id,
            func.count(TripUpdate.trip_update_id).label("update_count"),
        )
        .group_by(TripUpdate.route_id)
        .all()
    )

    # return list(result)
    return pd.DataFrame(result, columns=["route_id", "update_count"])

def get_trip_updates_by_date():
    result = (
        db.session.query(
            TripUpdate.date,
            func.count(TripUpdate.trip_update_id).label("update_count"),
        )
        .group_by(TripUpdate.date)
        .all()
    )

    # return list(result)
    return pd.DataFrame(result, columns=["date", "update_count"])


def get_alerts_by_date():
    result = (
        db.session.query(
            Alert.date,
            func.count(Alert.alert_id).label("count_alerts"),
        )
        .group_by(Alert.date)
        .all()
    )

    return pd.DataFrame(result, columns=["date", "count_alerts"])



def get_alerts_by_causes():
    result = (
        db.session.query(
            Alert.cause,
            func.count(Alert.alert_id).label("count_alerts"),
        )
        .group_by(Alert.cause)
        .all()
    )

    return pd.DataFrame(result, columns=["cause", "count_alerts"])


def get_alerts_by_effects():
    result = (
        db.session.query(
            Alert.effect,
            func.count(Alert.alert_id).label("count_alerts"),
        )
        .group_by(Alert.effect)
        .all()
    )
    return pd.DataFrame(result, columns=["effect", "count_alerts"])


def addAlertInfoToSheet(writer, sheet_name):

    alerts_by_date = get_alerts_by_date()
    alerts_by_effects = get_alerts_by_effects()
    alerts_by_causes = get_alerts_by_causes()
    alerts_by_routes = get_alerts_by_route()
    alerts_by_trips = get_alerts_by_trips()
    alerts_by_stops = get_alerts_by_stop()
    list_tables = {
        "Number of alerts by date issued": alerts_by_date,
        "Number of alerts by cause": alerts_by_causes,
        "Number of alerts by effects": alerts_by_effects,
        "Number of alerts by routes": alerts_by_routes,
        "Number of alerts by stops": alerts_by_stops,
        "Number of alerts by trips": alerts_by_trips,
    }
    write_dataframes_to_sheet(writer, sheet_name, list_tables)

def write_dataframes_to_sheet(writer, sheet_name, df_dict):
    worksheet = writer.book.get_sheet_by_name(sheet_name)
    current_row = 1
    for name, df in df_dict.items():
        numRows = len(df) + 2
        cell_header = worksheet.cell(row=current_row, column=1)
        cell_header.value = name
        df.to_excel(writer, sheet_name=sheet_name, index=False,header=True, startrow=current_row )
        current_row += numRows +1



def addTripUpdateInfoToSheet(writer, sheet_name):

    trip_updates_dates = get_trip_updates_by_date()
    trip_updates_routes = get_trip_updates_by_routes()
    trip_updates_trips = get_trip_updates_by_trips()
    trip_updates_stops = get_trip_updates_by_stops()
    list_tables = {
        "Number of trip updates by date issued": trip_updates_dates,
        "Number of trip updates by route": trip_updates_routes,
        "Number of trip updates by trips": trip_updates_trips,
        "Number of trip updates by stops": trip_updates_stops,
    }

    write_dataframes_to_sheet(writer, sheet_name, list_tables)


def create_service_excel(filename):
    writer = pd.ExcelWriter(filename, engine="openpyxl")
    try :
        workbook = writer.book
        workbook.create_sheet("Alerts")
        workbook.create_sheet("TripUpdates")
        addAlertInfoToSheet(writer, "Alerts")
        addTripUpdateInfoToSheet(writer, "TripUpdates")
        writer.close()
    except Exception as e:
        raise e
